from selenium import webdriver
from selenium.webdriver.common.by import By as by
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.edge.service import Service
from selenium.webdriver import EdgeOptions
import time, logging
import pandas as pd
from random import uniform
from itertools import zip_longest
from openpyxl import workbook, load_workbook



def Information_Collection(file_name, Category, Url): # main function containing all the action function and returns all the data
    web = '---' # website name
    website = f"{Url}" #website Url 
    driver = webdriver.Edge() # initializes driver with headless mode
    driver.get(website) # driver gets the website
    driver.maximize_window() # maximizes the window
    time.sleep(2)
    def button_Clicks(): # fuction for all the buttons to be clicked before scraping like pop-up window and scrolls through the page so all the data are loaded properly before scraping
        try:
            modal_close_button = driver.find_element(by.XPATH, "//div[@id = 'dsModalHeading']//button[@data-testid = 'modal-close-button']") # finds the close button for signing in pop up window
            modal_close_button.click() # clicks the close button for pop up 
            language_popup_window_close_button = driver.find_element(by.XPATH, "//div[@id = 'popup-container']//div[@id = 'popup-sticky-close']//button")
            language_popup_window_close_button.click()
        except:
            logging.error(f"Error scraping website: {e} in popupWindowClick Block")
        scrollable_element = driver.find_element(by.TAG_NAME, 'body') # finds the body of the websites html code
        scrollable_element.send_keys(Keys.END) # scrolls to the end of the webpage
        scrollable_element.send_keys(Keys.PAGE_UP)# clicks page up once so the middle part of the webpage is loaded properly
        try:
            buttons = driver.find_elements(by.XPATH, "//div[@class= 'sor-refresh-container']//button[@class = 'dsButtonTransparent_3oUqN dsA11yDrawerButton_AibMb']") # looks for the data area that is hidden in button click area
            for button in buttons:
                if str(button.get_attribute('aria-expanded')) == "false": # checks if the data area is hidden or not. if yes clicks the expand button else passes
                    button.click()
                else:
                    continue
        except Exception as e:
            logging.error(f"Error scraping website: {e} in colourAndSizeWindowClick Block")
        try:
            Dimension_box_display = driver.find_element(by.XPATH, "//div[@class = 'dimension-container']//h2//button[@type='button']").get_attribute('aria-expanded') # finds the dimension drop down html element
            if Dimension_box_display == "false": # checks if the datas are hidden or shown
                 driver.find_element(by.XPATH, "//div[@class = 'dimension-container']//h2//button[@type='button']").click() # clicks the dimension expand button
            time.sleep(2)
        except Exception as e:
            logging.error(f"Error scraping website: {e} in dimensionClick Block")

    


    def colourscraper(): # function that handles colour window and scrapes colour information
        all_colours = ""  # Initialize all_colours as an empty string
        try: # for if colour options are opened already when the website is loaded 
            colour_option_container = product_details_box.find_element(by.XPATH, ".//div[@class= 'sor-refresh-container']") # locates the information section of html
            colour_option = WebDriverWait(colour_option_container, 10).until(EC.presence_of_element_located((by.XPATH, ".//div[contains(@class, 'material-extra-style')]//div[@data-testid= 'content-test-id']//div[@data-attribute = 'fabric-content']")))
            colours = colour_option.find_elements(by.XPATH, ".//div[@class = 'quickViewContainer_1RHmQ']")
            for colour in colours:
                all_colours += f"{colour.find_element(by.XPATH, ".//label").text} || "
        except:
            try: # if colour options are down way slided
                colour_option = WebDriverWait(colour_option_container, 10).until(EC.presence_of_element_located((by.XPATH, ".//div[@class = 'material-old']//div[@data-attribute = 'fabric-content']")))
                colours = colour_option.find_elements(by.XPATH, ".//div[@class = 'quickViewContainer_1RHmQ']")
                for colour in colours:
                    all_colours += f"{colour.find_element(by.XPATH, ".//label").text} || "
            except:
                try: #if the colour options are right way slided
                    colour_container = product_details_box.find_element(by.XPATH, '//div[@class = "sor-refresh-container"]')
                    colour_box = colour_container.find_element(by.XPATH, ".//div[contains(@class, 'material-extra-style')]//div[@data-testid = 'content-test-id']")
                    colour_unite = colour_box.find_element(by.XPATH, ".//div[@data-attribute= 'material-container']")
                    colour_fields = colour_unite.find_elements(by.XPATH, ".//fieldset[@id = 'fabric-legend']")
                    for colour_field in colour_fields:  
                        colours_field_boxs = colour_field.find_elements(by.XPATH, ".//div[@class = 'fieldset-children']")
                        for colour_field_box in colours_field_boxs:
                            colours = colour_field_box.find_elements(by.XPATH, ".//div[@class = 'quickViewContainer_1RHmQ']")
                            for colour in colours: # iterates through all the colour options and appends them in the colour string
                                all_colours += f"{colour.find_element(by.XPATH, ".//label").text} || "
                except:
                    try: # for mainly leather options or with no colour option exsisting only one colour
                        colour_list = product_details_box.find_element(by.XPATH, ".//div[@class = 'custom-details-container custom-details-top']")
                        colour_list_item = colour_list.find_element(by.XPATH, ".//div[@class= 'custom-details-options']//button")
                        all_colours += f"{colour_list_item.text}"
                    except: # if no colour detail is provided
                        all_colours += 'Visit Product Link or Colour not Available'
                        print(all_colours)
        return all_colours
    



    def Photo_information(): # photo information scraper function
        photo_container = driver.find_element(by.XPATH, "//div[@class = 'mid-col gallery-col']") #locates the right or photo details containing section
        photos= photo_container.find_elements(by.XPATH,".//div[@style= 'width:100%']//button[@class= 'dsButtonTransparent_3U0k0 carouselButton_3dv9l']")
        main_image_url1 = '' # string intialization for photo links to be appended
        main_image_url2 = ''
        other_images = ''
        
        # checks the photo location in total list count. 1st photo is always thumbneil photo so that has been appended to other photos for cleaner photo option
        for i in range(len(photos)):
            if len(photos)>=4:
                if i == 0:
                    thumb_image = photos[i].find_element(by.XPATH, ".//img").get_attribute('src')
                if i == 2: # 2 because 0 is always the thumbneil photo and 1 is mostly dimension details photo is appened in main_image_url 1 and 2

                        main_image_url1 = photos[i].find_element(by.XPATH, ".//img").get_attribute('src')
                elif i==3: # 3 to take the then photo
                    try:
                        main_image_url2 =  photos[i].find_element(by.XPATH, ".//img").get_attribute('src')
                    except Exception as e:
                        logging.error(f"Error scraping website: {e} in photo block")
                else:
                    try:
                        other_images = other_images + f"{photos[i].find_element(by.XPATH, ".//img").get_attribute('src')}, "
                    except:
                        continue


            elif len(photos)==3: # as not more than 3 photo is available so thumbneil photo and other non dimensional photo is appened in main_image_url 1 and 2
                if i == 0:
                    main_image_url1 = photos[i].find_element(by.XPATH, ".//img").get_attribute('src')

                elif i==2:
                        main_image_url2 =  photos[i].find_element(by.XPATH, ".//img").get_attribute('src')
                else:
                        other_images = photos[i].find_element(by.XPATH, ".//img").get_attribute('src')



            elif len(photos)==2: # as less then 3 is abailable so we are using thumbneil photo in main photo and dimensional or non dimensional photo in 2nd main image url 
                        main_image_url1 = photos[0].find_element(by.XPATH, ".//img").get_attribute('src')
                        main_image_url2 = photos[1].find_element(by.XPATH, ".//img").get_attribute('src')
                        other_images = 'other image not available'

            elif len(photos)==1:
                main_image_url1 = photos[0].find_element(by.XPATH, ".//img").get_attribute('src')
                main_image_url2 = 'image2 not available'
                other_images = 'other image not available'

        return thumb_image, main_image_url1, main_image_url2, other_images




    def Details_Collector(): # function for other infomation scraping
        try:
          subtitle = product_details_box.find_element(by.XPATH, ".//div[@class = 'product-subtitle']").text # gets subtitle in the right column 
        except:
             subtitle = 'Subtitle not available'
        # print(subtitle) # for debugging
        try:
            price = product_details_box.find_element(by.XPATH,".//span[@class = 'salePrice']").text # gets sale price in the right column and regular price is not taken
        except:
            try:
                price = product_details_box.find_element(by.XPATH,".//span[@class = 'regPrice']").text # gets regular price if there are no sale price available
            except:
                price = 'Price not available'
        # print(price) # for debugging
        try:
            product_description = product_details_box.find_element(by.XPATH, ".//div[@class = 'details-description']").text # appends the whole product description
        except:
            product_description = 'Product description not available'
        product_hightlight = "" # initisalizes the highlight string
        try:
            product_hightlights = product_details_box.find_elements(by.XPATH, "//ul[@class = 'details-list']/li") # gets the list of highlights as they are not inputed in a paragaraph
            for highlight in product_hightlights: # iterates to all the element in highlights
               product_hightlight = product_hightlight + f"{highlight.text} || " # expands the string with the listed hightlights as they are listed in different lists
        except:
            product_hightlight = 'Product highlights not available'
        # print(product_hightlight) # for debugging
        try:
            rating_drawer = WebDriverWait(driver, 10).until(EC.presence_of_element_located((by.XPATH, "//div[@data-id= 'reviews-drawer']"))) # waiting upto 10 seconds to load the all the information in review box
        except:
            pass
        # print('rating') # for debugging
        try:
            rating = rating_drawer.find_element(by.XPATH, ".//div[@class = 'bv_rating_content2']//div[@class = 'bv-rnr__sc-157rd1w-1 fGWRcM']").text # gets the rating in review box
        except:
              rating = 'Product Not Rated'
        # print(rating)  # for debugging
        try:
            review_count = rating_drawer.find_element(by.XPATH, ".//section[@id='bv-reviews-overall-ratings-container']//div[@class = 'bv-rnr__sc-157rd1w-2 bPZJVP']").text # gets the amount of reviews that product got
        except:
            review_count = 'Not Available'
        # print(review_count) # for debugging
        product_id = website.split('/')[-1] # takes only the product id in the product url as it is the last element in the product url
        # print(product_id) # for debugging
        try:
            sku_number = product_details_box.find_element(by.XPATH, "//div[@class = 'right-col gallery-col']//div[@class = 'product-review-sku']//div").text # gets the sku number from right information box
        except:
            sku_number = 'Not Available'
        # print(sku_number) # for debugging
        try:
            dimension_container = product_details_box.find_element(by.XPATH, ".//div[@class = 'dimension-container ']") # get the dimension details from the list
            dimensions = dimension_container.find_elements(by.XPATH, ".//div[@class = 'dimensions-item']")
            product_dimension_details = '' # dimension string
            for dimension in dimensions: # iterates through all the dimension indexes  
               product_dimension_details = product_dimension_details + f"{dimension.text} ||" # appends all dimension details in dimension list
        except:
            product_dimension_details = 'Product dimension details not available'
        # print(product_dimension_details) # for debugging
        product_code = f"{web}-{product_id}" # takes website name and product id and joins them as "websitename"-(Product Id)
        try:
            item_size = size.find_element(by.XPATH, ".//span//button").text #appends the size of the product
        except:
            item_size = 'Fixed size'
        # print(item_size)
        category = Category
        product_type = driver.find_element(by.XPATH, "//div[@data-id = 'main-product-section']//ol[@class ='breadcrumb-list']//li[3]").text
        
        # print(web, subtitle, price, product_id, category, product_code, sku_number, item_size, rating, review_count, product_dimension_details, product_description, product_hightlight) # for debugging
        return  subtitle, price, product_id, category, product_type, product_code,  sku_number, item_size, rating, review_count, product_dimension_details, product_description, product_hightlight
    


    def appender(): # function to append row of data column wise 
        ws.append([Rating, web, website, Category, Product_type, file_name, Subtitle, Thumb_image, image1,  image2, other_images, Price, Product_id, Product_Code, Sku_number, Item_Size, colour_options,  Review_count, Product_Dimension_Details,  Product_Description, Product_Hightlight])                   
        wb.save('Product Information.xlsx') # saves the updated excel file after each time is appended with a new row



    product_details_box = driver.find_element(by.XPATH, "//div[@class = 'right-col gallery-col']") # locates the right or details containing column
    product_sizes = product_details_box.find_elements(by.XPATH, ".//div[@class= 'sor-refresh-container']//div[@class = 'dsA11yDrawerWrap_17Umw']//div[@class = 'fieldset-children']//div[@class = 'quickViewContainer_1RHmQ']") # checks if there are more then 1 size is available . if yes then creates 2 different entries for 2 different sizes
    if (len(product_sizes))>1:
        for size in product_sizes:         
            button_Clicks() # buttons are clicked as needed
            Thumb_image, image1, image2, other_images=Photo_information() # gets the images as returned value of photo scraper function
            colour_options = colourscraper() # take the colour string as output
            Subtitle, Price, Product_id, Category, Product_type, Product_Code, Sku_number, Item_Size, Rating, Review_count, Product_Dimension_Details, Product_Description, Product_Hightlight = Details_Collector() # defines the data accordingly with returned data
            appender() # appends the row of data of product
    else:
        button_Clicks() # clicks buttons
        Thumb_image,image1, image2, other_images=Photo_information()
        colour_options = colourscraper()
        Subtitle, Price, Product_id, Category, Product_type, Product_Code, Sku_number, Item_Size, Rating, Review_count, Product_Dimension_Details, Product_Description, Product_Hightlight = Details_Collector() # defines the data accordingly with returned data
        appender() # appends data



df = pd.read_excel("Product_List.xlsx") #reads the excel file with url to products, their title and thumbneil photo



# Remove duplicate rows
df.drop_duplicates(inplace=True)
# Save the updated DataFrame to a new Excel file
df.to_excel('Product_List.xlsx', index=False)


#loads the excel file where the scraped file will be added
wb = load_workbook("Product Information.xlsx") 
ws = wb.active # selects active sheet
Coloumn1 = "Title" # selects the first column of the excel file
Coloumn2 = "Product Category" # selects 2nd column
Coloumn3 = "Price" # selects 3rd column
Coloumn4= 'Product URL' # selects 4th column
# selects the columns for the dedicated data type
ws['A1'] = 'Rating'
ws['B1'] = 'Web'
ws['C1'] = 'Website'
ws['D1'] = 'Category'
ws['E1'] = 'Product Type'
ws['F1'] = 'File Name'
ws['G1'] = 'Subtitle'
ws['H1'] = 'Thumb Image'
ws['I1'] = 'Image 1'
ws['J1'] = 'Image 2'
ws['K1'] = 'Other Images'
ws['L1'] = 'Price'
ws['M1'] = 'Product ID'
ws['N1'] = 'Product Code'
ws['O1'] = 'Sku Number'
ws['P1'] = 'Item Size'
ws['Q1'] = 'Sofa Colours'
ws['R1'] = 'Review Count'
ws['S1'] = 'Product Dimension Details'
ws['T1'] = 'Product Description'
ws['U1'] = 'Product Highlight'
#iterates through the cell and row values in input excel file
for index, row in df.iterrows():
    Title = row[Coloumn1]
    Product_Category = row[Coloumn2]
    Product_URL = row[Coloumn4]
    time.sleep(uniform(0,3)) # randomizes the time between 0 to 3 seconds before sending request to the website
    try:
        Information_Collection(Title, Product_Category, Product_URL) # main function that takes product link, product name and product thumbneil photo as input from the product details excel file
    except Exception as e:
        logging.error(f"Error scraping website: {e} in information_collection function")
